import serial
import time
from openpyxl import load_workbook
# INPUT -- INPUT -- INPUT -- INPUT
#----------------------------------
filename = "SWAtest.xlsx"
#----------------------------------

# initializing the Arduino serial connection
Arduino = serial.Serial(port='COM3',baudrate=9600,timeout=1)
time.sleep(1)

# initializing the Excel Sheet
wb = load_workbook(filename)
ws = wb.active

t = 0

T1index = 2
T2index = 2
T3index = 2
T4index = 2
T5index = 2
ws['A1'].value = "time [s]"
ws['B1'].value = "T1 [C]"
ws['C1'].value = "T2 [C]"
ws['D1'].value = "T3 [C]"
ws['E1'].value = "T4 [C]"
ws['F1'].value = "T5 [C}"
wb.save(filename)

time.sleep(5)

while True:
    rawdata = Arduino.readline()
    data = rawdata.decode('utf-8')
    parsedata = data.split()
    for word in parsedata:
        try:
            if word == "Sensor1":
                index1 = parsedata.index('Sensor1')
                temp1 = parsedata[index1 + 1]
                print(f"Sensor 1 temp is {temp1} C")

                ws['B'+str(T1index)].value = float(temp1)
                T1index += 1
                wb.save(filename)

            if word == "Sensor2":
                index2 = parsedata.index('Sensor2')
                temp2 = parsedata[index2 + 1]
                print(f"Sensor 2 temp is {temp2} C")

                ws['C' + str(T2index)].value = float(temp2)
                T2index += 1
                wb.save(filename)

            if word == "Sensor3":
                index3 = parsedata.index('Sensor3')
                temp3 = parsedata[index3 + 1]
                print(f"Sensor 3 temp is {temp3} C")

                ws['D' + str(T3index)].value = float(temp3)
                T3index += 1
                wb.save(filename)

            if word == "Sensor4":
                index4 = parsedata.index('Sensor4')
                temp4 = parsedata[index4 + 1]
                print(f"Sensor 4 temp is {temp4} C")

                ws['E' + str(T4index)].value = float(temp4)
                T4index += 1
                wb.save(filename)

            if word == "Sensor5":
                index5 = parsedata.index('Sensor5')
                temp5 = parsedata[index5 + 1]
                print(f"Sensor 5 temp is {temp5} C")

                ws['F' + str(T5index)].value = float(temp5)
                ws['A' + str(T5index)].value = t
                t += 5
                T5index += 1
                wb.save(filename)
        except IndexError:
            pass